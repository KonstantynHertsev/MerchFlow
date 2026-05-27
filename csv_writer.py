"""
Export listings to upload-tool format.

Supported targets:
  lazy_merch    — LazyMerch desktop uploader  (.xlsx)
  flying_upload — Flying Upload               (.csv)
  merch_titans  — Merch Titans                (.csv)
"""

from io import BytesIO
from pathlib import Path

import config

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _colors(profile: dict) -> list[str]:
    c = profile.get("colors", config.DEFAULT_COLORS)
    return (c + [""] * 4)[:4]


def _base(listing: dict, profile: dict) -> dict:
    colors = _colors(profile)
    return {
        "image":       listing.get("_image_file", ""),
        "title":       listing.get("title", ""),
        "brand":       profile.get("brand", config.DEFAULT_BRAND),
        "bullet_1":    listing.get("bullet_1", ""),
        "bullet_2":    listing.get("bullet_2", ""),
        "bullet_3":    listing.get("bullet_3", ""),
        "bullet_4":    listing.get("bullet_4", ""),
        "bullet_5":    listing.get("bullet_5", ""),
        "description": listing.get("description", ""),
        "keywords":    listing.get("keywords", ""),
        "department":  profile.get("department", config.DEFAULT_DEPARTMENT),
        "fit_type":    profile.get("fit_type", config.DEFAULT_FIT),
        "color_1":     colors[0],
        "color_2":     colors[1],
        "color_3":     colors[2],
        "color_4":     colors[3],
        "price":       profile.get("price", config.DEFAULT_PRICE),
    }


_DEPT_TO_FIT = {
    "mens":    "Men",
    "womens":  "Women",
    "unisex":  "Men, Women",
    "youth":   "Youth",
    "girls":   "Girls",
    "boys":    "Youth",
}


# ---------------------------------------------------------------------------
# LazyMerch — XLSX
# Exact column headers from MBA.xlsx template
# ---------------------------------------------------------------------------

_LM_HEADERS = [
    "DesignPath ( ; seperated if you want to upload diffrent product types with who required other design files. If it's not set but required LazyMerch will autoresize it)",
    "Lazy Template",
    "Fit Type",
    "Preferred Language for translation",
    "Title (EN)", "Brand (EN)", "Bullet 1 (EN)", "Bullet 2 (EN)", "Description (EN)",
    "Title (DE)", "Brand (DE)", "Bullet 1 (DE)", "Bullet 2 (DE)", "Description (DE)",
    "Title (FR)", "Brand (FR)", "Bullet 1 (FR)", "Bullet 2 (FR)", "Description (FR)",
    "Title (IT)", "Brand (IT)", "Bullet 1 (IT)", "Bullet 2 (IT)", "Description (IT)",
    "Title (ES)", "Brand (ES)", "Bullet 1 (ES)", "Bullet 2 (ES)", "Description (ES)",
    "Title (JP)", "Brand (JP)", "Bullet 1 (JP)", "Bullet 2 (JP)", "Description (JP)",
    "Mba Update ID (just if you want to Update your existing Products)",
    'CustomArgs (, seperated - "back" place the Design on the backsite)',
]


def _build_lazy_merch_xlsx(listings: list[dict], profile: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "new"

    # Header row styling (match LazyMerch green)
    header_fill = PatternFill("solid", fgColor="92D050")
    header_font = Font(bold=True)

    for col, header in enumerate(_LM_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    fit_type = _DEPT_TO_FIT.get(
        profile.get("department", config.DEFAULT_DEPARTMENT), "Men"
    )

    for listing in listings:
        if listing.get("_error"):
            continue
        b = _base(listing, profile)
        row = [
            b["image"],   # DesignPath — user fills full local path
            "",           # Lazy Template — configured in LazyMerch
            fit_type,
            "EN",
            b["title"], b["brand"], b["bullet_1"], b["bullet_2"], b["description"],
            "", "", "", "", "",  # DE
            "", "", "", "", "",  # FR
            "", "", "", "", "",  # IT
            "", "", "", "", "",  # ES
            "", "", "", "", "",  # JP
            "",  # Mba Update ID
            "",  # CustomArgs
        ]
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Merch Titans — XLSX
# Exact column headers from layout.xlsx template
# ---------------------------------------------------------------------------

_MT_HEADERS = [
    "Image Path",
    "Title",
    "Description",
    "Tags",
    "Primary Tag for TeePublic",
    "Amazon Merch Brand",
    "Amazon Merch Bullet #1",
    "Amazon Merch Bullet #2",
]


def _build_merch_titans_xlsx(listings: list[dict], profile: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merch Titans Automation"

    header_font = Font(bold=True)
    for col, header in enumerate(_MT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font

    for listing in listings:
        if listing.get("_error"):
            continue
        b = _base(listing, profile)
        ws.append([
            b["image"],        # Image Path
            b["title"],        # Title
            b["description"],  # Description
            b["keywords"],     # Tags
            "",                # Primary Tag for TeePublic
            b["brand"],        # Amazon Merch Brand
            b["bullet_1"],     # Amazon Merch Bullet #1
            b["bullet_2"],     # Amazon Merch Bullet #2
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Flying Upload — XLSX
# Based on FlyingUploadMBA.xlsx (standard, EN only)
# ---------------------------------------------------------------------------

_FU_HEADERS = [
    "Image Path", "Input Language", "Title", "Description", "Tags",
    "Type", "Color", None, "Brand", "Bullet Points 1", "Bullet Points 2",
    "Color1", "Color2", "Color3", "Color4", "Color5",
    "Color6", "Color7", "Color8", "Color9", "Color10",
    "Product", "Marketplace",
    "Price US", "Price GB", "Price DE", "Price FR", "Price IT", "Price ES", "Price JP",
    "Print", "Draft", "Auto Translate", None,
    "Collection", "Category", "Background Color (Hex)",
]

_DEPT_TO_TYPE = {
    "mens":   "men",
    "womens": "women",
    "unisex": "men, women",
    "youth":  "youth",
    "girls":  "girls",
}


def _build_flying_upload_xlsx(listings: list[dict], profile: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flying Upload POD"

    header_font = Font(bold=True)
    for col, header in enumerate(_FU_HEADERS, start=1):
        if header:
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

    dept = profile.get("department", config.DEFAULT_DEPARTMENT)
    prod_type = _DEPT_TO_TYPE.get(dept, "men")
    colors = _colors(profile)
    # Pad to 10 colors
    colors10 = (colors + [""] * 10)[:10]
    price = profile.get("price", config.DEFAULT_PRICE)

    for listing in listings:
        if listing.get("_error"):
            continue
        b = _base(listing, profile)
        ws.append([
            b["image"],       # Image Path
            "EN",             # Input Language
            b["title"],       # Title
            b["description"], # Description
            b["keywords"],    # Tags
            prod_type,        # Type
            "",               # Color (primary — user fills)
            "",               # empty column
            b["brand"],       # Brand
            b["bullet_1"],    # Bullet Points 1
            b["bullet_2"],    # Bullet Points 2
            *colors10,        # Color1–Color10
            "Standard t-shirt",  # Product
            "US",             # Marketplace
            price,            # Price US
            "", "", "", "", "", "",  # Price GB/DE/FR/IT/ES/JP
            "front",          # Print
            "yes",            # Draft
            "",               # Auto Translate
            "",               # empty column
            "",               # Collection
            "",               # Category
            "",               # Background Color (Hex)
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Format registry
# ---------------------------------------------------------------------------

FORMATS = {
    "lazy_merch":   {"label": "LazyMerch",     "ext": "xlsx"},
    "flying_upload": {"label": "Flying Upload", "ext": "xlsx"},
    "merch_titans": {"label": "Merch Titans",   "ext": "xlsx"},
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_output(
    listings: list[dict],
    profile: dict | None = None,
    software: str = "lazy_merch",
) -> tuple[bytes, str]:
    """Returns (file_bytes, extension)."""
    profile = profile or {}
    if software == "lazy_merch":
        return _build_lazy_merch_xlsx(listings, profile), "xlsx"
    if software == "merch_titans":
        return _build_merch_titans_xlsx(listings, profile), "xlsx"
    if software == "flying_upload":
        return _build_flying_upload_xlsx(listings, profile), "xlsx"
    raise ValueError(f"Unknown software: {software}")


def write_csv(
    listings: list[dict],
    output_path: str | Path,
    profile: dict | None = None,
    software: str = "lazy_merch",
) -> Path:
    output_path = Path(output_path)
    data, _ = build_output(listings, profile or {}, software)
    output_path.write_bytes(data)
    return output_path
